import streamlit as st
import pandas as pd
import numpy as np
import re, io, json, zipfile, base64, requests
from copy import copy, deepcopy
from pathlib import Path
from PIL import Image
from openpyxl import load_workbook
import cv2
from rapidocr_onnxruntime import RapidOCR

st.set_page_config(page_title="三品牌採購系統", page_icon="🧾", layout="wide")

BRAND_OPTIONS = {
    "AN": "AN｜8碼主貨號",
    "AG": "AG｜13碼完整貨號",
    "JB": "JB｜8碼主貨號",
}
brand = st.sidebar.radio(
    "選擇品牌",
    list(BRAND_OPTIONS),
    format_func=lambda x: BRAND_OPTIONS[x],
    key="selected_brand",
)

# 切換品牌時清除畫面暫存，避免把上一品牌的照片或輸出帶過來。
if st.session_state.get("active_brand") != brand:
    st.session_state["active_brand"] = brand
    st.session_state["ocr_df"] = None
    st.session_state["generated_zip"] = None
    st.session_state.pop("editor", None)
    st.session_state.pop("images", None)
    for state_key in list(st.session_state.keys()):
        if str(state_key).startswith("AG_summary_"):
            st.session_state.pop(state_key, None)

st.title(f"🧾 {brand} 採購系統")
if brand == "AN":
    st.caption("商品照片 OCR → 資料複查 → 8碼主貨號 → 採購單＋商品基本資料")
elif brand == "AG":
    st.caption("商品照片 OCR → 資料複查 → 13碼完整貨號 → 採購單＋商品基本資料")
else:
    st.caption("商品照片 OCR → 資料複查 → 8碼主貨號／10碼商品型號 → 採購單＋商品基本資料")

# -------------------------
# 基本函式
# -------------------------
def clean(v):
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v)).upper()

def normalize_color(v):
    x = str(v or "").strip()
    return x[:-1] if x.endswith("色") else x

def read_excel_bytes(uploaded_file):
    return uploaded_file.getvalue()

def read_lookup(data, kind):
    raw = pd.read_excel(io.BytesIO(data), header=None, dtype=str).fillna("")
    result = {}
    for _, row in raw.iterrows():
        vals = [str(x).strip() for x in row.tolist() if str(x).strip()]
        if len(vals) < 2:
            continue
        a, b = vals[0], vals[1]
        joined = clean(a + b)
        if kind == "vendor":
            if "廠商代碼" in joined or "廠商名稱" in joined:
                continue
            if re.fullmatch(r"(?:\d{3,}|[A-Z]+\d+)", clean(a)):
                result[clean(b)] = a
            elif re.fullmatch(r"(?:\d{3,}|[A-Z]+\d+)", clean(b)):
                result[clean(a)] = b
        else:
            if re.fullmatch(r"\d+", clean(a)):
                result[clean(a)] = b
            elif re.fullmatch(r"\d+", clean(b)):
                result[clean(b)] = a
    return result

def read_color_lookup(data):
    raw = pd.read_excel(io.BytesIO(data), header=None, dtype=str).fillna("")
    cmap = {}
    for _, r in raw.iterrows():
        vals = [str(x).strip() for x in r.tolist() if str(x).strip()]
        if len(vals) >= 2 and "代號" not in "".join(vals):
            a, b = vals[0], vals[1]
            if re.fullmatch(r"\d+", a):
                cmap[normalize_color(b)] = a.zfill(2)
            elif re.fullmatch(r"\d+", b):
                cmap[normalize_color(a)] = b.zfill(2)
    return cmap

CATEGORY_PRODUCT_CODE = {
    "上衣":"1","短洋":"2","長洋":"3","褲子":"4",
    "裙子":"5","外套":"6","配件":"7","套裝":"8"
}

@st.cache_resource(show_spinner="第一次啟動 OCR 需要載入免費模型，請稍候…")
def get_reader():
    return RapidOCR()

def label_value(text, labels, pattern):
    for label in labels:
        m = re.search(re.escape(label) + r"\s*[:：]?\s*" + pattern, text, re.I)
        if m:
            return m.group(1).strip()
    return ""

def center(box):
    return (int(sum(p[0] for p in box)/4), int(sum(p[1] for p in box)/4))

def local_brightness(image, box):
    xs=[int(p[0]) for p in box]; ys=[int(p[1]) for p in box]
    pad=5
    x1=max(0,min(xs)-pad); x2=min(image.shape[1],max(xs)+pad)
    y1=max(0,min(ys)-pad); y2=min(image.shape[0],max(ys)+pad)
    crop=image[y1:y2,x1:x2]
    if not crop.size:
        return 0
    return float(np.median(cv2.cvtColor(crop,cv2.COLOR_BGR2GRAY)))

def normalize_ocr_text_for_variants(text):
    """
    顏色/尺碼/數量專用正規化。
    注意：這裡故意不使用類別/原編的簡繁正規化，避免互相影響。
    """
    s = str(text or "")
    s = s.replace("：", ":").replace("；", ";").replace("，", ",")
    s = s.replace("／", "/").replace("、", "/")
    s = re.sub(r"\s+", "", s)

    # 只處理顏色常見簡繁，不碰「類別」「原編」等欄位
    color_replacements = {
        "黑色": "黑", "黑": "黑",
        "白色": "白", "白": "白",
        "卡其色": "卡其", "卡其": "卡其",
        "咖啡色": "咖啡", "咖啡": "咖啡",
        "棕色": "棕", "棕": "棕",
        "灰色": "灰", "灰": "灰",
        "藍色": "藍", "蓝色": "藍", "蓝": "藍",
        "紅色": "紅", "红色": "紅", "红": "紅",
        "綠色": "綠", "绿色": "綠", "绿": "綠",
        "粉色": "粉", "粉": "粉",
        "米色": "米", "米": "米",
        "杏色": "杏", "杏": "杏",
        "駝色": "駝", "驼色": "駝", "驼": "駝",
        "紫色": "紫", "紫": "紫",
        "黃色": "黃", "黄色": "黃", "黄": "黃",
        "橘色": "橘", "橙色": "橘",
        "深灰色": "深灰", "深灰": "深灰",
        "淺灰色": "淺灰", "浅灰色": "淺灰", "浅灰": "淺灰",
        "深藍色": "深藍", "深蓝色": "深藍", "深蓝": "深藍",
        "淺藍色": "淺藍", "浅蓝色": "淺藍", "浅蓝": "淺藍",
        "米白色": "米白", "米白": "米白",
        "藏青色": "藏青", "藏青": "藏青",
    }
    for a, b in sorted(color_replacements.items(), key=lambda x: len(x[0]), reverse=True):
        s = s.replace(a, b)

    return s


def normalize_qty_token(token):
    """
    數量專用 OCR 修正：
    O/Q/D -> 0
    I/L/|/! -> 1
    Z -> 2
    S -> 5
    B -> 8
    只保留數字。
    """
    s = str(token or "").upper()
    trans = {
        "O":"0", "Q":"0", "D":"0",
        "I":"1", "L":"1", "|":"1", "!":"1",
        "Z":"2", "S":"5", "B":"8",
    }
    s = "".join(trans.get(ch, ch) for ch in s)
    s = re.sub(r"[^0-9]", "", s)
    return int(s) if s else None


def parse_size_qty_segment(segment):
    """
    尺寸+數量專用解析。
    尺寸固定接受 XS/S/M/L/XL/2XL/3XL/F，
    數量則只取尺寸後面緊鄰的 OCR 數字/常見誤判字元。
    """
    segment = str(segment or "").upper()
    results = []
    pos = 0

    size_tokens = ["XXXL", "3XL", "XXL", "2XL", "XL", "XS", "FREE", "S", "M", "L", "F"]

    while pos < len(segment):
        matched = False

        for token in size_tokens:
            if segment.startswith(token, pos):
                j = pos + len(token)

                # 容忍分隔符
                while j < len(segment) and segment[j] in ":xX*×-/":
                    j += 1

                # 數量最多抓 4 字元，避免吃到下一個尺寸
                k = j
                qty_chars = []
                while k < len(segment) and len(qty_chars) < 4:
                    ch = segment[k]
                    # 遇到下一個尺寸字母就停止
                    if ch in "SMLFX" and qty_chars:
                        break
                    if ch.isdigit() or ch in "OQDILZSB|!":
                        qty_chars.append(ch)
                        k += 1
                    else:
                        break

                qty = normalize_qty_token("".join(qty_chars))
                if qty is not None:
                    size = "F" if token == "FREE" else token
                    results.append((size, qty))
                    pos = k
                    matched = True
                    break

        if matched:
            continue

        pos += 1

    return results


def quantity_diagnostics(variants):
    """
    數量異常檢查（安全版）：
    - None / 空值：只提醒，不中斷 OCR
    - 0 或負數
    - > 999
    - 同一顏色下某尺寸數量與其他尺寸差異過大
    """
    warnings = []
    by_color = {}

    for item in variants:
        if len(item) < 3:
            continue

        color, size, qty = item[0], item[1], item[2]
        by_color.setdefault(color, []).append((size, qty))

        if qty is None or (isinstance(qty, str) and not qty.strip()):
            warnings.append(f"{color}-{size} 數量未辨識，請確認")
            continue

        try:
            q = int(float(qty))
        except (TypeError, ValueError):
            warnings.append(f"{color}-{size} 數量「{qty}」無法判讀，請確認")
            continue

        if q <= 0:
            warnings.append(f"{color}-{size} 數量異常：{q}")
        elif q > 999:
            warnings.append(f"{color}-{size} 數量疑似辨識錯誤：{q}")

    for color, rows in by_color.items():
        vals = []
        safe_rows = []

        for size, qty in rows:
            if qty is None or (isinstance(qty, str) and not qty.strip()):
                continue
            try:
                q = int(float(qty))
            except (TypeError, ValueError):
                continue
            if q > 0:
                vals.append(q)
                safe_rows.append((size, q))

        if len(vals) >= 3:
            sorted_vals = sorted(vals)
            med = sorted_vals[len(sorted_vals) // 2]

            if med > 0:
                for size, q in safe_rows:
                    if q >= med * 3 or q * 3 <= med:
                        warnings.append(
                            f"{color}-{size} 數量 {q} 與同色其他尺寸差異過大，請確認"
                        )

    return warnings


def parse_sizes_qty(text):
    """
    顏色 -> 尺寸 -> 數量 專用解析。
    支援：
      黑S40M40L40卡其S40M40L40
      黑 S40 M40 L40 / 卡其 S40 M40 L40
      黑:S40/M40/L40；卡其:S40/M40/L40
    """
    s = normalize_ocr_text_for_variants(text)

    color_names = [
        "深咖啡","淺咖啡","咖啡",
        "卡其",
        "深灰","淺灰","灰",
        "藏青","深藍","淺藍","藍",
        "米白","米",
        "粉","紅","綠","白","黑",
        "杏","駝","棕","紫","黃","橘"
    ]

    color_pat = "|".join(map(re.escape, sorted(color_names, key=len, reverse=True)))
    hits = list(re.finditer(color_pat, s))

    variants = []
    for i, hit in enumerate(hits):
        color = hit.group()
        seg_start = hit.end()
        seg_end = hits[i + 1].start() if i + 1 < len(hits) else len(s)
        segment = s[seg_start:seg_end]

        for size, qty in parse_size_qty_segment(segment):
            variants.append((color, size, qty))

    # 去除完全重複項目
    deduped = []
    seen = set()
    for item in variants:
        if item not in seen:
            seen.add(item)
            deduped.append(item)

    return deduped


def variant_parse_diagnostics(text, variants):
    """
    只檢查顏色/尺寸，不碰類別/原編正規化。
    """
    s = normalize_ocr_text_for_variants(text)

    known_colors = [
        "深咖啡","淺咖啡","咖啡","卡其",
        "深灰","淺灰","灰","藏青","深藍","淺藍","藍",
        "米白","米","粉","紅","綠","白","黑",
        "杏","駝","棕","紫","黃","橘"
    ]
    color_pat = "|".join(map(re.escape, sorted(known_colors, key=len, reverse=True)))
    hits = list(re.finditer(color_pat, s))

    parsed_by_color = {}
    for color, size, qty in variants:
        parsed_by_color.setdefault(color, set()).add(size)

    missing_colors = []
    missing_sizes = []

    for i, hit in enumerate(hits):
        color = hit.group()

        if color not in parsed_by_color:
            if color not in missing_colors:
                missing_colors.append(color)
            continue

        seg_end = hits[i + 1].start() if i + 1 < len(hits) else len(s)
        segment = s[hit.end():seg_end]
        expected_sizes = {size for size, qty in parse_size_qty_segment(segment)}

        for size in expected_sizes:
            if size not in parsed_by_color.get(color, set()):
                missing_sizes.append(f"{color}-{size}")

    return {
        "missing_colors": missing_colors,
        "missing_sizes": missing_sizes,
    }



def rapidocr_details(reader, image):
    """
    RapidOCR 結果轉成舊程式使用的 (box, text, confidence) 格式。
    """
    result, elapsed = reader(image)
    details = []
    if not result:
        return details

    for item in result:
        try:
            box, txt, score = item[0], item[1], item[2]
            details.append((box, str(txt), float(score)))
        except Exception:
            continue

    return details


def crop_text_region(image):
    """
    商品資訊通常集中在上半部，保留 65% 高度。
    """
    h, w = image.shape[:2]
    return image[:max(1, int(h * 0.65)), :]


def preprocess_text_region(region):
    """
    文字區適度放大 + 灰階對比強化。
    """
    h, w = region.shape[:2]
    max_width = 1500

    if w > max_width:
        scale = max_width / float(w)
        region = cv2.resize(
            region,
            (max_width, max(1, int(h * scale))),
            interpolation=cv2.INTER_AREA
        )
    else:
        scale = min(1.8, max_width / max(1, w))
        if scale > 1.10:
            region = cv2.resize(
                region,
                None,
                fx=scale,
                fy=scale,
                interpolation=cv2.INTER_CUBIC
            )

    gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)

    return region, enhanced


def ocr_text_from_details(details):
    if not details:
        return ""

    ordered = sorted(
        details,
        key=lambda x: (center(x[0])[1], center(x[0])[0])
    )
    return " ".join(x[1] for x in ordered).replace("臺", "台")


def free_ocr_bundle(reader, image):
    """
    免費 RapidOCR 主流程：
    1. 原圖辨識一般欄位
    2. 上方文字區辨識顏色/尺碼/數量
    3. 對比強化文字區再辨識一次
    4. 選擇能拆出最多顏色/尺寸的結果
    """
    original_details = rapidocr_details(reader, image)
    original_text = ocr_text_from_details(original_details)

    region = crop_text_region(image)

    candidate_texts = []

    # 文字區原圖
    d1 = rapidocr_details(reader, region)
    t1 = ocr_text_from_details(d1)
    candidate_texts.append(t1)

    # 對比強化文字區
    _, enhanced = preprocess_text_region(region)
    d2 = rapidocr_details(reader, enhanced)
    t2 = ocr_text_from_details(d2)
    candidate_texts.append(t2)

    # 先用原圖做基準
    best_text = original_text
    best_variants = parse_sizes_qty(original_text)

    def _score(parsed):
        return (
            len(parsed),
            len({x[0] for x in parsed}) if parsed else 0,
            len({x[1] for x in parsed}) if parsed else 0,
        )

    best_score = _score(best_variants)

    for txt in candidate_texts:
        parsed = parse_sizes_qty(txt)
        score = _score(parsed)
        if score > best_score:
            best_text = txt
            best_variants = parsed
            best_score = score

    general_text = (original_text + " " + best_text).strip()

    return original_details, general_text, best_text, best_variants



def compact_ocr_text(text):
    s = str(text or "").replace("臺", "台").replace("：", ":")
    s = re.sub(r"\s+", "", s)

    replacements = {
        "类别": "類別",
        "类別": "類別",
        "分类": "分類",
        "品类": "品類",
        "类型": "類型",
        "原编": "原編",
        "原厂编号": "原廠編號",
        "原廠编号": "原廠編號",
        "原厂編號": "原廠編號",
        "款号": "款號",
        "型号": "型號",
        "货号": "貨號",
        "售价": "售價",
        "进价": "進價",
    }
    for a, b in replacements.items():
        s = s.replace(a, b)
    return s


def details_to_lines(details, y_tolerance=18):
    if not details:
        return []

    items = []
    for box, txt, conf in details:
        try:
            cx, cy = center(box)
            items.append((cy, cx, str(txt).strip()))
        except Exception:
            continue

    items.sort(key=lambda x: (x[0], x[1]))
    lines = []

    for cy, cx, txt in items:
        if not txt:
            continue
        if not lines or abs(cy - lines[-1]["y"]) > y_tolerance:
            lines.append({"y": cy, "items": [(cx, txt)]})
        else:
            lines[-1]["items"].append((cx, txt))
            lines[-1]["y"] = int((lines[-1]["y"] + cy) / 2)

    result = []
    for line in lines:
        line["items"].sort(key=lambda x: x[0])
        result.append(" ".join(t for _, t in line["items"]))
    return result


def normalize_category_candidate(value):
    x = clean(value)
    replacements = {"S":"5", "B":"8", "I":"1", "L":"1", "Z":"2"}
    x = replacements.get(x, x)
    m = re.search(r"[1-8]", x)
    return m.group(0) if m else ""


def extract_category_code_robust(all_text, details=None):
    compact = compact_ocr_text(all_text)
    labels = ["類別", "分類", "品類", "類型"]
    label_alt = "|".join(map(re.escape, labels))

    # 標準數字
    m = re.search(rf"(?:{label_alt})(?:代碼)?[:#\-]?([1-8])", compact, re.I)
    if m:
        return m.group(1)

    # OCR 常見誤判
    m = re.search(rf"(?:{label_alt})(?:代碼)?[:#\-]?([SBILZ])", compact, re.I)
    if m:
        fixed = normalize_category_candidate(m.group(1))
        if fixed:
            return fixed

    # 行級重組
    lines = details_to_lines(details or [])
    for i, line in enumerate(lines):
        cline = compact_ocr_text(line)
        if any(label in cline for label in labels):
            for label in labels:
                if label in cline:
                    tail = cline.split(label, 1)[1]
                    fixed = normalize_category_candidate(tail[:5])
                    if fixed:
                        return fixed
            if i + 1 < len(lines):
                fixed = normalize_category_candidate(compact_ocr_text(lines[i+1])[:3])
                if fixed:
                    return fixed

    # fuzzy fallback
    for pat in [
        r"(?:別|别)([1-8])",
        r"(?:別|别)([SBILZ])",
        r"(?:類|类)[^\dSBILZ]{0,2}([1-8SBILZ])",
    ]:
        m = re.search(pat, compact, re.I)
        if m:
            fixed = normalize_category_candidate(m.group(1))
            if fixed:
                return fixed

    return ""


def category_debug_snippet(all_text, details=None):
    compact = compact_ocr_text(all_text)
    keywords = ["類", "別", "别", "分", "品", "型"]
    lines = details_to_lines(details or [])
    hits = []
    for line in lines:
        c = compact_ocr_text(line)
        if any(k in c for k in keywords):
            hits.append(c)
    if hits:
        return "｜".join(hits[:4])
    return compact[:160]


def normalize_factory_no(value):
    x = str(value or "").strip().upper()
    x = re.sub(r"^[：:\-#]+", "", x)
    x = re.sub(r"[^A-Z0-9\-/]", "", x)
    return x


def extract_factory_no_robust(all_text, details=None, known_numbers=None):
    compact = compact_ocr_text(all_text)

    labels = ["原廠編號", "原廠編", "原編號", "原編", "款號", "款式號", "型號"]
    label_alt = "|".join(map(re.escape, labels))

    m = re.search(
        rf"(?:{label_alt})[:#\-]?([A-Z0-9][A-Z0-9\-/]{{2,19}})",
        compact,
        re.I
    )
    if m:
        value = normalize_factory_no(m.group(1))
        if 3 <= len(value) <= 20:
            return value

    lines = details_to_lines(details or [])
    for i, line in enumerate(lines):
        cline = compact_ocr_text(line)
        if any(label in cline for label in labels):
            for label in labels:
                if label in cline:
                    tail = cline.split(label, 1)[1]
                    val = normalize_factory_no(tail)
                    if len(val) >= 3:
                        return val
            if i + 1 < len(lines):
                nxt = normalize_factory_no(lines[i+1])
                if re.fullmatch(r"[A-Z0-9][A-Z0-9\-/]{2,19}", nxt):
                    return nxt

    excludes = {str(x) for x in (known_numbers or []) if str(x)}
    candidates = re.findall(r"(?<!\d)(\d{4,12})(?!\d)", compact)
    candidates = [x for x in candidates if x not in excludes]
    if candidates:
        candidates.sort(key=lambda x: (-len(x), compact.find(x)))
        return candidates[0]

    return ""

def parse_image(uploaded_image, vendor_map, category_map):
    raw = uploaded_image.getvalue()
    arr = np.frombuffer(raw, np.uint8)
    image = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    reader=get_reader()
    details, all_text, color_size_text, color_size_variants = free_ocr_bundle(reader, image)
    ordered=sorted(details,key=lambda x:(center(x[0])[1],center(x[0])[0]))

    vendor=""
    for key in sorted(vendor_map,key=len,reverse=True):
        if key and key in clean(all_text):
            vendor=key
            break
    cost=label_value(all_text,["成本","進價"],r"([0-9,]+)")
    price=label_value(all_text,["售價","定價"],r"([0-9,]+)")

    category_code=extract_category_code_robust(all_text, details)

    factory_no=extract_factory_no_robust(
        all_text,
        details,
        known_numbers=[
            cost.replace(",","") if cost else "",
            price.replace(",","") if price else "",
            category_code,
        ]
    )
    pm=re.search(r"(\d+(?:\.\d+)?)\s*%",all_text)
    percent=(pm.group(1)+"%") if pm else ""

    exclusions=["廠商","成本","進價","售價","定價","原編","原编","原廠編號","原厂编号","類別","类别","分類","分类"]
    summary_parts=[]
    for box,text,conf in ordered:
        t=text.strip()
        if not t or any(k in t for k in exclusions) or re.fullmatch(r"\d+(?:\.\d+)?%",t):
            continue
        if local_brightness(image,box)>=165 and not re.fullmatch(r"[A-Z]?\d+",t,re.I):
            summary_parts.append(t)
    summary="、".join(dict.fromkeys(summary_parts))

    variants=color_size_variants if color_size_variants else parse_sizes_qty(all_text)
    diagnostics=variant_parse_diagnostics(color_size_text or all_text, variants)

    field_warning_parts=[]
    if not factory_no:
        field_warning_parts.append("原廠編號未辨識")
    if not category_code:
        debug_text = category_debug_snippet(all_text, details)
        field_warning_parts.append(f"類別代碼未辨識（OCR片段：{debug_text}）")

    if not variants:
        variants=[("","",None)]

    warning_parts=[]
    if diagnostics.get("missing_colors"):
        warning_parts.append("可能漏顏色："+"、".join(diagnostics["missing_colors"]))
    if diagnostics.get("missing_sizes"):
        warning_parts.append("可能漏尺寸："+"、".join(diagnostics["missing_sizes"]))

    qty_warnings = quantity_diagnostics(variants)
    if qty_warnings:
        warning_parts.extend(qty_warnings)
    warning_parts.extend(field_warning_parts)
    parse_warning="；".join(warning_parts)

    rows=[]
    for color,size,qty in variants:
        rows.append({
            "來源照片": uploaded_image.name,
            "廠商": vendor,
            "廠商代碼": vendor_map.get(vendor,""),
            "原廠編號": factory_no,
            "貨號":"",
            "類別代碼": category_code,
            "類別": category_map.get(clean(category_code),""),
            "顏色":color,
            "尺寸":size,
            "摘要":summary,
            "進價":int(cost.replace(",","")) if cost else None,
            "售價":int(price.replace(",","")) if price else None,
            "數量":qty,
            "到貨日":"",
            "備註1":"",
            "備註2":percent,
            "OCR文字":all_text,
            "辨識警示":parse_warning
        })
    return rows

def update_vendor_map_from_table(df, vendor_map):
    updated=dict(vendor_map)
    for _, r in df.iterrows():
        name=clean(r.get("廠商",""))
        code=str(r.get("廠商代碼","") or "").strip()
        if name and code:
            updated[name]=code
    return updated

def vendor_map_to_excel(vendor_map):
    rows=[{"廠商代碼":code,"廠商名稱":name} for name,code in vendor_map.items()]
    out=io.BytesIO()
    pd.DataFrame(rows).to_excel(out,index=False)
    return out.getvalue()

def build_category_next_sequences(ledger):
    """
    依既有 assignments 自動反推各類別的下一個流水號。
    8碼貨號第2碼為類別碼，最後3碼為流水號。
    舊版只有全域 next_sequence 也沒關係，會從既有貨號自動轉成分類紀錄。
    """
    seq_map = dict(ledger.get("next_sequence_by_category", {}) or {})

    for info in (ledger.get("assignments", {}) or {}).values():
        base = str(info.get("base8", "") or "").strip()
        seq = info.get("sequence", None)

        if re.fullmatch(r"\d{8}", base):
            catcode = base[1]
            try:
                used_seq = int(seq) if seq is not None else int(base[-3:])
            except Exception:
                used_seq = int(base[-3:])
            seq_map[catcode] = max(int(seq_map.get(catcode, 1)), used_seq + 1)

    # 每個已知類別至少從 001 開始
    for catcode in CATEGORY_PRODUCT_CODE.values():
        seq_map.setdefault(str(catcode), 1)

    return seq_map


def assign_an_codes(df, ledger, custom_code, seq_mode, manual_starts=None):
    working = deepcopy(ledger)
    working.setdefault("assignments", {})
    working["next_sequence_by_category"] = build_category_next_sequences(working)

    base_by_key = {}
    statuses = []
    used_base8 = {
        str(v.get("base8", "")).strip()
        for v in working["assignments"].values()
        if str(v.get("base8", "")).strip()
    }

    # 手動模式也改成「每個類別自己的起始號」
    manual_starts = manual_starts or {}
    manual_current = {str(k): int(v) for k, v in manual_starts.items()}

    for (vendor_code, original), g in df.groupby(["廠商代碼", "原廠編號"], sort=False):
        key = f"{vendor_code}|{original}"
        old = working["assignments"].get(key)

        # 已建檔商品永遠沿用舊貨號
        if old:
            base = old["base8"]
            statuses.append(f"{original}：沿用 {base}")
        else:
            cat = str(g.iloc[0]["類別"]).strip()
            catcode = str(CATEGORY_PRODUCT_CODE.get(cat, ""))
            if not catcode:
                raise ValueError(f"類別「{cat}」沒有對應的貨號類別碼。")

            if seq_mode == "manual":
                seq = int(manual_current.get(
                    catcode,
                    working["next_sequence_by_category"].get(catcode, 1)
                ))
            else:
                seq = int(working["next_sequence_by_category"].get(catcode, 1))

            if seq > 999:
                raise ValueError(f"類別「{cat}」流水號已超過 999。")

            base = f"6{catcode}{custom_code}{seq:03d}"
            if base in used_base8:
                raise ValueError(
                    f"貨號 {base} 已存在。請調整「{cat}」的起始流水號。"
                )

            working["assignments"][key] = {
                "base8": base,
                "sequence": seq,
                "category": cat,
                "category_code": catcode,
                "vendor_code": str(vendor_code),
                "original": str(original)
            }
            used_base8.add(base)

            # 只推進「本類別」自己的流水號
            next_seq = seq + 1
            working["next_sequence_by_category"][catcode] = max(
                int(working["next_sequence_by_category"].get(catcode, 1)),
                next_seq
            )

            if seq_mode == "manual":
                manual_current[catcode] = next_seq
                statuses.append(f"{original}：新編 {base}（{cat}｜自行設定）")
            else:
                statuses.append(f"{original}：新編 {base}（{cat}｜系統接續）")

        base_by_key[key] = base

    # 保留舊欄位相容性，但不再用它決定新貨號
    all_next = list(working["next_sequence_by_category"].values())
    working["next_sequence"] = max(all_next) if all_next else 1

    result = df.copy()
    result["貨號"] = result.apply(
        lambda r: base_by_key[f'{r["廠商代碼"]}|{r["原廠編號"]}'],
        axis=1
    )
    return result, working, base_by_key, statuses


def build_global_next_sequence(ledger, code_field, sequence_slice):
    """AG／JB 使用品牌各自獨立的4碼全域流水號。"""
    next_no = int(ledger.get("next_sequence", 1) or 1)
    for info in (ledger.get("assignments", {}) or {}).values():
        code = str(info.get(code_field, "") or "").strip()
        seq = info.get("sequence")
        try:
            used = int(seq) if seq is not None else int(code[sequence_slice])
            next_no = max(next_no, used + 1)
        except Exception:
            continue
    return next_no


def build_category_next_sequences_4(ledger, code_field):
    """由既有 AG／JB 貨號反推各類別自己的4碼下一號。"""
    seq_map = dict(ledger.get("next_sequence_by_category", {}) or {})
    for info in (ledger.get("assignments", {}) or {}).values():
        code = str(info.get(code_field, "") or "").strip()
        if not re.fullmatch(r"\d{8,11}", code):
            continue
        catcode = code[1]
        try:
            used_seq = int(info.get("sequence")) if info.get("sequence") is not None else int(code[2:6])
        except Exception:
            continue
        seq_map[catcode] = max(int(seq_map.get(catcode, 1)), used_seq + 1)
    for catcode in CATEGORY_PRODUCT_CODE.values():
        seq_map.setdefault(str(catcode), 1)
    return seq_map


def assign_ag_codes(df, ledger, season_digit, custom_code, seq_mode, manual_starts, color_map):
    """
    AG 13碼：季別1＋類別1＋流水4＋自填2＋售價去個位3＋顏色2。
    同款不同顏色共用流水號，只更換最後2碼顏色。
    """
    working = deepcopy(ledger)
    working.setdefault("assignments", {})
    working["next_sequence_by_category"] = build_category_next_sequences_4(working, "prefix11")
    prefix_by_key = {}
    statuses = []
    used = {
        str(v.get("prefix11", "") or "").strip()
        for v in working["assignments"].values()
        if str(v.get("prefix11", "") or "").strip()
    }
    used_sequences = {
        (code[1], int(code[2:6]))
        for code in used
        if re.fullmatch(r"\d{11}", code)
    }
    manual_current = {str(k): int(v) for k, v in (manual_starts or {}).items()}

    for (vendor_code, original), g in df.groupby(["廠商代碼", "原廠編號"], sort=False):
        key = f"{vendor_code}|{original}"
        old = working["assignments"].get(key)
        if old and re.fullmatch(r"\d{11}", str(old.get("prefix11", ""))):
            prefix11 = str(old["prefix11"])
            statuses.append(f"{original}：沿用前11碼 {prefix11}")
        else:
            cat = str(g.iloc[0]["類別"]).strip()
            catcode = str(CATEGORY_PRODUCT_CODE.get(cat, ""))
            if not catcode:
                raise ValueError(f"類別「{cat}」沒有對應的第2碼。")
            price = safe_int(g.iloc[0].get("售價"))
            if price is None or price < 0:
                raise ValueError(f"原廠編號 {original} 的售價不正確。")
            price_value = price // 10
            if price_value > 999:
                raise ValueError(f"售價 {price} 去掉個位數後超過3碼。")
            sequence = int(manual_current.get(catcode, working["next_sequence_by_category"].get(catcode, 1))) if seq_mode == "manual" else int(working["next_sequence_by_category"].get(catcode, 1))
            if sequence > 9999:
                raise ValueError("AG 流水號已超過9999。")
            if (catcode, sequence) in used_sequences:
                raise ValueError(f"AG「{cat}」流水號 {sequence:04d} 已被其他款式使用，請改用不同號碼。")
            prefix11 = f"{season_digit}{catcode}{sequence:04d}{custom_code}{price_value:03d}"
            if prefix11 in used:
                raise ValueError(f"AG 前11碼 {prefix11} 已存在，請調整起始流水號。")
            working["assignments"][key] = {
                "prefix11": prefix11,
                "sequence": sequence,
                "category": cat,
                "category_code": catcode,
                "vendor_code": str(vendor_code),
                "original": str(original),
                "price": price,
            }
            used.add(prefix11)
            used_sequences.add((catcode, sequence))
            next_seq = sequence + 1
            working["next_sequence_by_category"][catcode] = max(int(working["next_sequence_by_category"].get(catcode, 1)), next_seq)
            if seq_mode == "manual":
                manual_current[catcode] = next_seq
            statuses.append(f"{original}：新編前11碼 {prefix11}（{cat}）")
        prefix_by_key[key] = prefix11

    result = df.copy()
    def full_ag_code(row):
        cname = normalize_color(row.get("顏色", ""))
        ccode = str(color_map.get(cname, "") or "")
        if not re.fullmatch(r"\d{2}", ccode):
            raise ValueError(f"顏色「{cname}」找不到2碼代號。")
        key = f'{row["廠商代碼"]}|{row["原廠編號"]}'
        return prefix_by_key[key] + ccode
    result["貨號"] = result.apply(full_ag_code, axis=1)
    working["next_sequence"] = max(working["next_sequence_by_category"].values())
    return result, working, prefix_by_key, statuses


def assign_jb_codes(df, ledger, brand_code, custom_code, seq_mode, manual_sequences=None):
    """JB 8碼：自填品牌1＋類別1＋流水4＋自填2。"""
    working = deepcopy(ledger)
    working.setdefault("assignments", {})
    working["next_sequence_by_category"] = build_category_next_sequences_4(working, "base8")
    base_by_key = {}
    statuses = []
    used = {
        str(v.get("base8", "") or "").strip()
        for v in working["assignments"].values()
        if str(v.get("base8", "") or "").strip()
    }
    used_sequences = {
        (code[1], int(code[2:6]))
        for code in used
        if re.fullmatch(r"\d{8}", code)
    }
    manual_sequences = manual_sequences or {}
    for (vendor_code, original), g in df.groupby(["廠商代碼", "原廠編號"], sort=False):
        key = f"{vendor_code}|{original}"
        old = working["assignments"].get(key)
        if old and re.fullmatch(r"\d{8}", str(old.get("base8", ""))):
            base8 = str(old["base8"])
            statuses.append(f"{original}：沿用 {base8}")
        else:
            cat = str(g.iloc[0]["類別"]).strip()
            catcode = str(CATEGORY_PRODUCT_CODE.get(cat, ""))
            if not catcode:
                raise ValueError(f"類別「{cat}」沒有對應的第2碼。")
            sequence = int(manual_sequences.get(key, working["next_sequence_by_category"].get(catcode, 1))) if seq_mode == "manual" else int(working["next_sequence_by_category"].get(catcode, 1))
            if sequence > 9999:
                raise ValueError("JB 流水號已超過9999。")
            if (catcode, sequence) in used_sequences:
                raise ValueError(f"JB「{cat}」流水號 {sequence:04d} 已被其他款式使用，請改用不同號碼。")
            base8 = f"{brand_code}{catcode}{sequence:04d}{custom_code}"
            if base8 in used:
                raise ValueError(f"JB 貨號 {base8} 已存在，請調整起始流水號。")
            working["assignments"][key] = {
                "base8": base8,
                "sequence": sequence,
                "category": cat,
                "category_code": catcode,
                "brand_code": str(brand_code),
                "vendor_code": str(vendor_code),
                "original": str(original),
            }
            used.add(base8)
            used_sequences.add((catcode, sequence))
            statuses.append(f"{original}：新編 {base8}（{cat}）")
            working["next_sequence_by_category"][catcode] = max(int(working["next_sequence_by_category"].get(catcode, 1)), sequence + 1)
        base_by_key[key] = base8
    result = df.copy()
    result["貨號"] = result.apply(
        lambda r: base_by_key[f'{r["廠商代碼"]}|{r["原廠編號"]}'], axis=1
    )
    working["next_sequence"] = max(working["next_sequence_by_category"].values())
    return result, working, base_by_key, statuses

def safe_int(v, default=None):
    if v is None:
        return default
    try:
        if pd.isna(v):
            return default
    except Exception:
        pass
    s = str(v).strip()
    if not s:
        return default
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return default


def create_purchase_workbooks(df, template_bytes, brand_name, brand_code, optional_date, optional_page):
    outputs={}
    for vendor,group in df.groupby("廠商",sort=False):
        wb=load_workbook(io.BytesIO(template_bytes))
        ws=wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.active
        headers={
            str(ws.cell(2,c).value).strip():c
            for c in range(1,ws.max_column+1) if ws.cell(2,c).value
        }
        ws["A1"]=f'廠商: {vendor}                                  廠代: {group.iloc[0]["廠商代碼"]}                                   品牌代碼: {brand_code}'
        if optional_page:
            ws.cell(1,ws.max_column,f"頁數:{optional_page}")
        for n,(_,r) in enumerate(group.iterrows(),3):
            for col in range(1,ws.max_column+1):
                src,dst=ws.cell(2,col),ws.cell(n,col)
                dst._style=copy(src._style)
            vals={
                "日期":optional_date,
                "原廠編號":str(r["原廠編號"]),
                "貨號":r["貨號"],
                "類別":r["類別"],
                "顏色":r["顏色"],
                "尺寸":r["尺寸"],
                "摘要":r["摘要"],
                "進價":safe_int(r["進價"], ""),
                "售價":safe_int(r["售價"], ""),
                "數量":safe_int(r["數量"], ""),
                "到貨日":"",
                "備註1":"",
                "備註2":r["備註2"]
            }
            for h,v in vals.items():
                if h in headers:
                    ws.cell(n,headers[h],v)
        bio=io.BytesIO()
        wb.save(bio)
        outputs[f'{brand_name}採購單-{group.iloc[0]["廠商代碼"]}.xlsx']=bio.getvalue()
    return outputs

def create_master_workbook(df, template_bytes, code_by_key, color_map, brand_name, brand_code, season, optional_date):
    target=['商品型號','品名規格','供應廠商','品牌編號','建議售價','起始進價','最後進價',
            '特價','類別1','類別2','類別3','類別4','類別5','尺碼代號','季別','建檔日期',
            '備註1','原廠編號','材質四']
    records=[]
    missing_colors=[]
    for (vendor_code,original),g in df.groupby(["廠商代碼","原廠編號"],sort=False):
        base=code_by_key[f"{vendor_code}|{original}"]
        first=g.iloc[0]
        for color,cg in g.groupby("顏色",sort=False):
            cname=normalize_color(color)
            ccode=color_map.get(cname,"")
            if not re.fullmatch(r"\d{2}",str(ccode)):
                missing_colors.append(cname)
                continue
            sizes={str(x).strip().upper() for x in cg["尺寸"] if str(x).strip()}
            sizecode="4" if sizes=={"F"} else "5"
            if brand_name == "AG":
                product_code = base + str(ccode)
            else:
                product_code = base + str(ccode)
            summary_number = str(first.get("摘要", "") or "").strip()
            records.append({
                "商品型號":product_code,
                "品名規格":product_code+cname if brand_name == "AG" else base+cname,
                "供應廠商":vendor_code,
                "品牌編號":brand_code,
                "建議售價":safe_int(first["售價"], ""),
                "起始進價":safe_int(first["進價"], ""),
                "最後進價":safe_int(first["進價"], ""),
                "特價":"",
                "類別1":"08",
                "類別2":summary_number if brand_name == "AG" else "",
                "類別3":summary_number if brand_name == "AG" else "",
                "類別4":summary_number if brand_name == "AG" else "",
                "類別5":str(ccode),"尺碼代號":sizecode,
                "季別":season,"建檔日期":optional_date,
                "備註1":"","原廠編號":original,"材質四":first["備註2"],
                "數量":sum(safe_int(v, 0) or 0 for v in cg["數量"])
            })
    if missing_colors:
        raise ValueError("以下顏色找不到 2 碼代號："+"、".join(sorted(set(missing_colors))))
    result=pd.DataFrame(records,columns=target)
    wb=load_workbook(io.BytesIO(template_bytes))
    ws=wb.active
    heads=[str(c.value or "").strip() for c in ws[1]]
    if ws.max_row>1:
        ws.delete_rows(2,ws.max_row-1)
    for rn,rec in enumerate(result.to_dict("records"),2):
        for cn,h in enumerate(heads,1):
            cell=ws.cell(rn,cn,rec.get(h,""))
            cell._style=copy(ws.cell(1,cn)._style)
            if h in ["商品型號","品名規格","供應廠商","品牌編號","類別1","類別2","類別3","類別4","類別5","尺碼代號","原廠編號"]:
                cell.number_format="@"
    bio=io.BytesIO()
    wb.save(bio)
    name=f'{brand_name}商品基本資料-{df.iloc[0]["廠商代碼"]}.xlsx'
    return name,bio.getvalue(),result


# -------------------------
# GitHub 永久保存（data 分支）
# -------------------------
GITHUB_OWNER = "WAWAPIG0303"
GITHUB_REPO = "an-ag-jb-purchase-system"
DATA_BRANCH = "data"
AN_CONFIG_PATHS = {
    "vendor": "data/廠商代碼.xlsx",
    "category": "data/類別.xlsx",
    "color": "data/類別5.xlsx",
    "purchase_template": "data/採購單空白範本.xlsx",
    "master_template": "data/商品基本資料空白範本.xlsx",
    "ledger": "data/貨號流水記錄_三碼001.json",
}
BRAND_CONFIG_PATHS = {
    "AN": AN_CONFIG_PATHS,
    "AG": {
        "vendor": "data/AG/廠商代碼.xlsx",
        "category": "data/類別.xlsx",
        "color": "data/AG/類別5.xlsx",
        "purchase_template": "data/AG/採購單空白範本.xlsx",
        "master_template": "data/AG/商品基本資料空白範本.xlsx",
        "ledger": "data/AG/貨號流水記錄_四碼0001.json",
    },
    "JB": {
        "vendor": "data/JB/廠商代碼.xlsx",
        "category": "data/類別.xlsx",
        "color": "data/JB/類別5.xlsx",
        "purchase_template": "data/JB/採購單空白範本.xlsx",
        "master_template": "data/JB/商品基本資料空白範本.xlsx",
        "ledger": "data/JB/貨號流水記錄_四碼0001.json",
    },
}
CONFIG_PATHS = BRAND_CONFIG_PATHS[brand]

def get_token():
    try:
        return st.secrets["GITHUB_TOKEN"]
    except Exception:
        return ""

def gh_headers():
    return {
        "Authorization": f"Bearer {get_token()}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

def gh_api(path, method="GET", payload=None):
    url=f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}{path}"
    return requests.request(method,url,headers=gh_headers(),json=payload,timeout=30)

def ensure_data_branch():
    r=gh_api(f"/git/ref/heads/{DATA_BRANCH}")
    if r.status_code==200:
        return
    main=gh_api("/git/ref/heads/main")
    if main.status_code!=200:
        raise RuntimeError("無法讀取 main 分支，請檢查 GITHUB_TOKEN 權限。")
    sha=main.json()["object"]["sha"]
    new=gh_api("/git/refs",method="POST",payload={"ref":f"refs/heads/{DATA_BRANCH}","sha":sha})
    if new.status_code not in (200,201,422):
        raise RuntimeError(f"無法建立 data 分支：{new.status_code}")

def gh_read_bytes(path):
    r=gh_api(f"/contents/{path}?ref={DATA_BRANCH}")
    if r.status_code==404:
        return None
    if r.status_code!=200:
        raise RuntimeError(f"讀取 {path} 失敗：{r.status_code}")
    return base64.b64decode(r.json()["content"])

def gh_write_bytes(path, content, message):
    cur=gh_api(f"/contents/{path}?ref={DATA_BRANCH}")
    payload={"message":message,"content":base64.b64encode(content).decode("ascii"),"branch":DATA_BRANCH}
    if cur.status_code==200:
        payload["sha"]=cur.json()["sha"]
    elif cur.status_code!=404:
        raise RuntimeError(f"檢查 {path} 失敗：{cur.status_code}")
    r=gh_api(f"/contents/{path}",method="PUT",payload=payload)
    if r.status_code not in (200,201):
        raise RuntimeError(f"儲存 {path} 失敗：{r.status_code} {r.text[:160]}")

def github_ready():
    if not get_token():
        return False,"尚未設定 GITHUB_TOKEN"
    r=gh_api("")
    if r.status_code!=200:
        return False,"GITHUB_TOKEN 無法讀取此儲存庫"
    try:
        ensure_data_branch()
    except Exception as e:
        return False,str(e)
    return True,""

# -------------------------
# Session state
# -------------------------
for k,v in {"ocr_df":None,"generated_zip":None}.items():
    if k not in st.session_state:
        st.session_state[k]=v

# -------------------------
# 永久保存連線與初始化
# -------------------------
ready, ready_msg = github_ready()
if not ready:
    st.error("尚未完成永久保存設定")
    st.code(ready_msg)
    st.info("請先到 Streamlit → Manage app → Settings → Secrets，加入 GITHUB_TOKEN。")
    st.stop()

config_bytes={}
missing=[]
for k,p in CONFIG_PATHS.items():
    b=gh_read_bytes(p)
    if b is None:
        missing.append(k)
    else:
        config_bytes[k]=b

fixed_keys=["vendor","category","color","purchase_template","master_template"]

# AG／JB 第一次使用時，先複製 AN 現有固定資料到各自目錄。
# 複製後是獨立檔案，之後更新任一品牌不會影響其他品牌。
if brand != "AN" and any(k in missing for k in fixed_keys):
    for k in fixed_keys:
        if k not in missing:
            continue
        source = gh_read_bytes(AN_CONFIG_PATHS[k])
        if source is not None:
            gh_write_bytes(CONFIG_PATHS[k], source, f"初始化 {brand} {k}")
            config_bytes[k] = source
            missing.remove(k)

if any(k in missing for k in fixed_keys):
    st.header("🔧 第一次初始化")
    st.info("只做這一次。完成後，以後開啟網頁只需要上傳商品照片。")
    c1,c2,c3=st.columns(3)
    with c1:
        f_vendor=st.file_uploader("廠商代碼.xlsx",type=["xlsx"])
        f_category=st.file_uploader("類別.xlsx",type=["xlsx"])
    with c2:
        f_color=st.file_uploader("類別5.xlsx（顏色代號）",type=["xlsx"])
        f_purchase=st.file_uploader("採購單空白範本",type=["xlsx"])
    with c3:
        f_master=st.file_uploader("商品基本資料空白範本",type=["xlsx"])
    if st.button("💾 儲存固定設定",type="primary"):
        if not all([f_vendor,f_category,f_color,f_purchase,f_master]):
            st.error("請先上傳全部 5 個固定檔案。")
        else:
            with st.spinner("正在永久保存…"):
                gh_write_bytes(CONFIG_PATHS["vendor"],f_vendor.getvalue(),f"初始化 {brand} 廠商代碼")
                gh_write_bytes(CONFIG_PATHS["category"],f_category.getvalue(),f"初始化 {brand} 類別")
                gh_write_bytes(CONFIG_PATHS["color"],f_color.getvalue(),f"初始化 {brand} 顏色代號")
                gh_write_bytes(CONFIG_PATHS["purchase_template"],f_purchase.getvalue(),f"初始化 {brand} 採購單範本")
                gh_write_bytes(CONFIG_PATHS["master_template"],f_master.getvalue(),f"初始化 {brand} 商品基本資料範本")
                if gh_read_bytes(CONFIG_PATHS["ledger"]) is None:
                    default=json.dumps({"next_sequence":1,"assignments":{}},ensure_ascii=False,indent=2).encode("utf-8")
                    gh_write_bytes(CONFIG_PATHS["ledger"],default,"初始化貨號流水")
            st.success("固定設定已永久保存。請重新整理頁面。")
    st.stop()

if "ledger" not in config_bytes:
    default=json.dumps({"next_sequence":1,"assignments":{}},ensure_ascii=False,indent=2).encode("utf-8")
    gh_write_bytes(CONFIG_PATHS["ledger"],default,"初始化貨號流水")
    config_bytes["ledger"]=default

vendor_map=read_lookup(config_bytes["vendor"],"vendor")
category_map=read_lookup(config_bytes["category"],"category")
color_map=read_color_lookup(config_bytes["color"])
ledger=json.loads(config_bytes["ledger"].decode("utf-8"))
ledger.setdefault("next_sequence",1)
ledger.setdefault("assignments",{})
if brand == "AN":
    ledger["next_sequence_by_category"] = build_category_next_sequences(ledger)
else:
    code_field = "prefix11" if brand == "AG" else "base8"
    ledger["next_sequence_by_category"] = build_category_next_sequences_4(ledger, code_field)
if brand in ("AN", "AG", "JB"):
    width = 3 if brand == "AN" else 4
    seq_summary = "｜".join(
        f'{cat} {int(ledger["next_sequence_by_category"].get(str(code),1)):0{width}d}'
        for cat, code in CATEGORY_PRODUCT_CODE.items()
    )
    st.success(f"{brand} 設定已自動載入｜各類別下一號：{seq_summary}")

with st.expander("⚙️ 管理固定設定（平常不用開）"):
    replace_kind=st.selectbox("要更新的資料",[
        "廠商代碼.xlsx","類別.xlsx","類別5.xlsx","採購單空白範本.xlsx","商品基本資料空白範本.xlsx"
    ],key=f"{brand}_replace_kind")
    replace_file=st.file_uploader("選擇新版檔案",type=["xlsx"],key=f"{brand}_replace_config")
    replace_map={"廠商代碼.xlsx":"vendor","類別.xlsx":"category","類別5.xlsx":"color",
                 "採購單空白範本.xlsx":"purchase_template","商品基本資料空白範本.xlsx":"master_template"}
    if st.button("更新固定設定",disabled=replace_file is None,key=f"{brand}_replace_button"):
        key=replace_map[replace_kind]
        gh_write_bytes(CONFIG_PATHS[key],replace_file.getvalue(),f"更新 {replace_kind}")
        st.success("更新完成，重新整理後生效。")


def rebuild_rows_from_variant_text(edited_df):
    """
    依每張照片的「顏色尺寸原文」重新建立顏色/尺寸/數量。
    不重新跑 OCR，因此穩定、低資源。
    """
    rebuilt = []
    errors = []

    for source, group in edited_df.groupby("來源照片", sort=False):
        base = group.iloc[0].to_dict()
        raw = str(base.get("顏色尺寸原文", "") or "").strip()

        if not raw:
            errors.append(f"{source}：顏色尺寸原文是空白")
            rebuilt.extend(group.to_dict("records"))
            continue

        parsed = parse_sizes_qty(raw)
        if not parsed:
            errors.append(f"{source}：無法從「{raw}」拆出顏色尺寸")
            rebuilt.extend(group.to_dict("records"))
            continue

        for color, size, qty in parsed:
            row = dict(base)
            row["顏色"] = color
            row["尺寸"] = size
            row["數量"] = qty
            row["辨識警示"] = ""
            rebuilt.append(row)

    return pd.DataFrame(rebuilt), errors



def append_manual_variant(df, source_value, color, size, qty):
    """
    人工新增顏色/尺寸/數量：
    複製同商品既有完整資料，只覆蓋顏色、尺寸、數量。
    避免 data_editor 新增空白列後因廠商/原編/類別為空而在輸出 groupby 時被忽略。
    """
    if df is None or df.empty:
        raise ValueError("目前沒有可新增的商品資料。")

    work = df.copy()

    # source_value 格式：來源照片｜原廠編號
    source_photo = source_value.split("｜", 1)[0].strip()
    original = source_value.split("｜", 1)[1].strip() if "｜" in source_value else ""

    mask = work["來源照片"].astype(str).eq(source_photo)
    if original and "原廠編號" in work.columns:
        mask &= work["原廠編號"].astype(str).eq(original)

    candidates = work[mask]
    if candidates.empty:
        raise ValueError("找不到要新增資料的商品。")

    base = candidates.iloc[0].to_dict()
    base["顏色"] = str(color or "").strip()
    base["尺寸"] = str(size or "").strip().upper()
    base["數量"] = safe_int(qty, None)
    base["辨識警示"] = ""
    base["資料來源"] = "人工新增"

    # 顏色尺寸原文不再拿來覆蓋人工新增內容
    if "顏色尺寸原文" in base:
        base["顏色尺寸原文"] = str(base.get("顏色尺寸原文", "") or "")

    return pd.concat([work, pd.DataFrame([base])], ignore_index=True)


def mark_manual_edits_as_final(df):
    """
    將目前 data_editor 畫面視為最終確認資料。
    後續產生 Excel 一律使用這份，不再重新 OCR / 重新拆分。
    """
    out = df.copy()
    if "資料來源" not in out.columns:
        out["資料來源"] = ""
    out["資料來源"] = out["資料來源"].replace("", "人工確認")
    return out

# -------------------------
# 日常操作：只上傳照片
# -------------------------
st.header("① 上傳商品照片")
images=st.file_uploader("可一次選多張 JPG / PNG / WEBP",type=["jpg","jpeg","png","webp"],accept_multiple_files=True,key="images")

if st.button("🔍 開始 OCR 辨識",type="primary",disabled=not images):
    try:
        for state_key in list(st.session_state.keys()):
            if str(state_key).startswith("AG_summary_"):
                st.session_state.pop(state_key, None)
        rows=[]; progress=st.progress(0)
        for i,img in enumerate(images):
            rows.extend(parse_image(img,vendor_map,category_map))
            progress.progress((i+1)/len(images))
        st.session_state.ocr_df=pd.DataFrame(rows)
        st.success(f"完成，共辨識 {len(images)} 張照片。")
    except Exception as e:
        st.error(f"OCR 失敗：{e}")

if st.session_state.ocr_df is not None:
    st.header("② 確認／修改資料")
    st.info("可以直接點表格修改。新廠商請填入『廠商名稱＋廠商代碼』，產生後會自動永久保存。")
    edited=st.data_editor(
        st.session_state.ocr_df.drop(columns=["OCR文字"],errors="ignore"),
        num_rows="dynamic",
        use_container_width=True,
        key="editor"
    )
    edited=edited.copy()

    # 目前畫面上的人工修改就是最終資料來源
    edited = mark_manual_edits_as_final(edited)

    st.info(
        "人工修改優先：你在這個表格修改的顏色、尺寸、數量，"
        "以及下方「新增顏色／尺寸／數量」加入的資料，"
        "都會直接用於最後 Excel，不會再被 OCR 覆蓋。"
    )

    with st.expander("➕ 新增顏色／尺寸／數量", expanded=False):
        product_options = []
        if not edited.empty and all(c in edited.columns for c in ["來源照片","原廠編號"]):
            seen = set()
            for _, r in edited.iterrows():
                label = f'{str(r.get("來源照片","")).strip()}｜{str(r.get("原廠編號","")).strip()}'
                if label not in seen:
                    seen.add(label)
                    product_options.append(label)

        manual_product = st.selectbox(
            "選擇要新增的商品",
            product_options,
            key=f"{brand}_manual_variant_product"
        ) if product_options else None

        mc1, mc2, mc3 = st.columns(3)
        with mc1:
            manual_color = st.text_input("顏色", key=f"{brand}_manual_variant_color", placeholder="例如：卡其")
        with mc2:
            manual_size = st.text_input("尺寸", key=f"{brand}_manual_variant_size", placeholder="例如：M")
        with mc3:
            manual_qty = st.number_input(
                "數量",
                min_value=0,
                max_value=9999,
                value=0,
                step=1,
                key=f"{brand}_manual_variant_qty"
            )

        if st.button("➕ 加入這筆資料", type="primary", key=f"{brand}_add_manual_variant"):
            try:
                if not manual_product:
                    raise ValueError("請先選擇商品。")
                if not str(manual_color).strip():
                    raise ValueError("請填顏色。")
                if not str(manual_size).strip():
                    raise ValueError("請填尺寸。")
                if int(manual_qty) <= 0:
                    raise ValueError("數量必須大於 0。")

                base_df = edited.copy()
                new_df = append_manual_variant(
                    base_df,
                    manual_product,
                    manual_color,
                    manual_size,
                    int(manual_qty)
                )
                st.session_state.ocr_df = new_df
                # 清掉 data_editor widget 狀態，讓新列完整顯示
                if "editor" in st.session_state:
                    del st.session_state["editor"]
                st.success("已加入人工資料，這筆會帶入最後 Excel。")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    st.caption(
        "「依顏色尺寸原文重新拆分」只在你想整批重建 OCR 顏色資料時使用。"
        "若你已經人工修改／新增完成，就不要再按重新拆分，以免重新建立列。"
    )

    if st.button("🔄 依顏色尺寸原文重新拆分"):
        rebuilt_df, rebuild_errors = rebuild_rows_from_variant_text(edited)
        st.session_state.ocr_df = rebuilt_df
        if rebuild_errors:
            st.warning("；".join(rebuild_errors))
        else:
            st.success("顏色、尺碼、數量已重新拆分。")
        st.rerun()

    if "辨識警示" in edited.columns:
        warnings=[
            str(x).strip()
            for x in edited["辨識警示"].fillna("").tolist()
            if str(x).strip()
        ]
        if warnings:
            st.warning(
                "⚠️ 系統偵測到顏色／尺寸可能漏辨識，請先確認表格後再產生採購單。\n\n"
                + "\n".join("• "+w for w in sorted(set(warnings)))
            )
    for idx,row in edited.iterrows():
        vname=clean(row.get("廠商",""))
        if vname in vendor_map and not str(row.get("廠商代碼","") or "").strip():
            edited.at[idx,"廠商代碼"]=vendor_map[vname]
        ccode=clean(row.get("類別代碼",""))
        if ccode in category_map:
            edited.at[idx,"類別"]=category_map[ccode]

    # AG 摘要逐款輸入一次，所有顏色／尺寸共用同一數字。
    if brand == "AG":
        st.subheader("AG 摘要數字（同款所有顏色共用）")
        for n, ((vendor_code, original), indexes) in enumerate(
            edited.groupby(["廠商代碼", "原廠編號"], sort=False).groups.items(), 1
        ):
            idx_list = list(indexes)
            value = st.text_input(
                f"原廠編號 {original} 的摘要數字",
                value="",
                key=f"AG_summary_{n}_{vendor_code}_{original}",
            ).strip()
            edited.loc[idx_list, "摘要"] = value

    a,b,c,d=st.columns(4)
    if brand == "AN":
        with a: brand_code=st.text_input("品牌代碼",value="6",key="AN_brand_code")
        with b: season=st.text_input("季別",key="AN_season")
        with c: custom_code=st.text_input("8碼貨號第3～5碼",max_chars=3,placeholder="例如 813",key="AN_custom")
    elif brand == "AG":
        with a: season=st.radio("季別／第1碼",["春夏","秋冬"],horizontal=True,key="AG_season")
        brand_code="5" if season=="春夏" else "9"
        with b: st.text_input("第1碼",value=brand_code,disabled=True,key="AG_brand_digit")
        with c: custom_code=st.text_input("13碼第7～8碼（整批共用）",max_chars=2,placeholder="例如 23",key="AG_custom")
    else:
        with a: brand_code=st.text_input("品牌代碼／第1碼",value="3",max_chars=1,key="JB_brand_digit")
        with b: season=st.text_input("季別",key="JB_season")
        with c: custom_code=st.text_input("8碼第7～8碼（整批共用）",max_chars=2,placeholder="例如 23",key="JB_custom")
    with d: optional_date=st.text_input("日期（選填）",placeholder="例如 8/24",key=f"{brand}_date")
    optional_page=st.text_input("頁數（選填）",key=f"{brand}_page")

    batch_categories = []
    manual_starts = {}
    jb_manual_sequences = {}
    manual_start = int(ledger.get("next_sequence", 1))

    if "類別" in edited.columns:
        for cat in edited["類別"].dropna().astype(str):
            cat = cat.strip()
            if cat in CATEGORY_PRODUCT_CODE and cat not in batch_categories:
                batch_categories.append(cat)

    if brand == "AN":
        st.subheader("商品流水號（依類別分開計算）")

        if batch_categories:
            cols = st.columns(min(4, len(batch_categories)))
            for i, cat in enumerate(batch_categories):
                catcode = str(CATEGORY_PRODUCT_CODE[cat])
                next_no = int(ledger["next_sequence_by_category"].get(catcode, 1))
                with cols[i % len(cols)]:
                    st.metric(f"{cat} 下一號", f"{next_no:03d}")
        else:
            st.caption("完成類別確認後，這裡會顯示各類別下一個流水號。")

        seq_label = st.radio(
            "編號方式",
            ["按照各類別系統紀錄接續", "自行設定各類別起始流水號"],
            horizontal=True,
            key="AN_seq_mode",
        )
        seq_mode = "system" if seq_label == "按照各類別系統紀錄接續" else "manual"
        if seq_mode == "manual":
            st.caption("只需設定本批有使用到的類別；不同類別可填不同起始號碼。")
            mcols = st.columns(min(4, max(1, len(batch_categories))))
            for i, cat in enumerate(batch_categories):
                catcode = str(CATEGORY_PRODUCT_CODE[cat])
                default_no = int(ledger["next_sequence_by_category"].get(catcode, 1))
                with mcols[i % len(mcols)]:
                    manual_starts[catcode] = st.number_input(
                        f"{cat} 起始流水號", min_value=1, max_value=999,
                        value=default_no, step=1, key=f"AN_manual_seq_{catcode}"
                    )
    elif brand == "AG":
        st.subheader("AG 款式流水號（依類別分開計算）")
        cols=st.columns(min(4,max(1,len(batch_categories))))
        for i,cat in enumerate(batch_categories):
            catcode=str(CATEGORY_PRODUCT_CODE[cat])
            with cols[i % len(cols)]: st.metric(f"{cat} 下一號",f'{int(ledger["next_sequence_by_category"].get(catcode,1)):04d}')
        seq_label = st.radio(
            "編號方式", ["按照各類別系統紀錄接續", "自行設定各類別起始流水號"],
            horizontal=True, key=f"{brand}_seq_mode"
        )
        seq_mode = "system" if seq_label == "按照各類別系統紀錄接續" else "manual"
        if seq_mode == "manual":
            mcols=st.columns(min(4,max(1,len(batch_categories))))
            for i,cat in enumerate(batch_categories):
                catcode=str(CATEGORY_PRODUCT_CODE[cat])
                with mcols[i % len(mcols)]:
                    manual_starts[catcode]=st.number_input(f"{cat} 起始流水號",min_value=1,max_value=9999,value=int(ledger["next_sequence_by_category"].get(catcode,1)),step=1,key=f"AG_manual_seq_{catcode}")
    else:
        st.subheader("JB 款式流水號（依類別分開計算）")
        cols=st.columns(min(4,max(1,len(batch_categories))))
        for i,cat in enumerate(batch_categories):
            catcode=str(CATEGORY_PRODUCT_CODE[cat])
            with cols[i % len(cols)]: st.metric(f"{cat} 下一號",f'{int(ledger["next_sequence_by_category"].get(catcode,1)):04d}')
        seq_label = st.radio(
            "編號方式", ["按照各類別系統紀錄接續", "逐款自行設定流水號"],
            horizontal=True, key="JB_seq_mode"
        )
        seq_mode = "system" if seq_label == "按照各類別系統紀錄接續" else "manual"
        if seq_mode == "manual":
            st.caption("每個不同款式輸入各自的4碼流水號；同款不同顏色會共用。")
            style_keys=[]
            for (vendor_code, original), _group in edited.groupby(["廠商代碼", "原廠編號"], sort=False):
                key=f"{vendor_code}|{original}"
                if key not in style_keys:
                    style_keys.append(key)
            cols=st.columns(min(3, max(1, len(style_keys))))
            category_offsets={}
            for i,key in enumerate(style_keys):
                vendor_code,original=key.split("|",1)
                style_group=edited[(edited["廠商代碼"].astype(str)==vendor_code)&(edited["原廠編號"].astype(str)==original)]
                cat=str(style_group.iloc[0]["類別"]).strip(); catcode=str(CATEGORY_PRODUCT_CODE.get(cat,""))
                old=ledger.get("assignments",{}).get(key,{})
                old_base=str(old.get("base8","") or "")
                offset=category_offsets.get(catcode,0)
                default_seq=int(old_base[2:6]) if re.fullmatch(r"\d{8}",old_base) else min(9999,int(ledger["next_sequence_by_category"].get(catcode,1))+offset)
                if not re.fullmatch(r"\d{8}",old_base): category_offsets[catcode]=offset+1
                with cols[i % len(cols)]:
                    jb_manual_sequences[key]=st.number_input(
                        f"{cat}｜款式 {original} 流水號", min_value=1, max_value=9999,
                        value=default_seq, step=1, key=f"JB_manual_seq_{vendor_code}_{original}"
                    )

    live_color_map=dict(color_map)
    missing_colors=[]
    if "顏色" in edited.columns:
        for x in edited["顏色"].dropna().astype(str):
            cname=normalize_color(x)
            if cname and cname not in live_color_map:
                missing_colors.append(cname)
    if missing_colors:
        st.warning("以下顏色尚無代號，請補上 2 碼：")
        for cname in sorted(set(missing_colors)):
            val=st.text_input(f"{cname} 的2碼顏色代號",max_chars=2,key=f"{brand}_color_{cname}")
            if re.fullmatch(r"\d{2}",val or ""):
                live_color_map[cname]=val

    if all(col in edited.columns for col in ["顏色","尺寸","數量"]):
        st.subheader("數量快速檢查")
        try:
            qty_check = (
                edited.groupby("顏色", dropna=False)["數量"]
                .sum()
                .reset_index()
                .rename(columns={"數量":"顏色總數量"})
            )
            st.dataframe(qty_check, use_container_width=True, hide_index=True)
        except Exception:
            pass

    st.header("③ 產生採購單")
    if st.button("✅ 產生採購單＋商品基本資料",type="primary"):
        try:
            required=["廠商","廠商代碼","原廠編號","類別代碼","類別","顏色","尺寸","進價","售價","數量","備註2"]
            if brand == "AG":
                required.append("摘要")
            problems=[]
            for i,row in edited.iterrows():
                miss=[col for col in required if pd.isna(row.get(col)) or str(row.get(col)).strip()==""]
                if miss: problems.append(f'第 {i+1} 列缺少：{"、".join(miss)}')
            if problems: raise ValueError("；".join(problems[:8]))
            if brand == "AN" and not re.fullmatch(r"\d{3}",custom_code or ""):
                raise ValueError("AN 8碼貨號第3～5碼必須輸入3位數字。")
            if brand in ("AG", "JB") and not re.fullmatch(r"\d{2}",custom_code or ""):
                raise ValueError(f"{brand} 第7～8碼必須輸入2位數字。")
            if brand == "JB" and not re.fullmatch(r"\d",brand_code or ""):
                raise ValueError("JB 品牌代碼／第1碼必須輸入1位數字。")
            # 人工確認資料優先：產生前固定使用目前畫面 edited
            confirmed_df = mark_manual_edits_as_final(edited)
            st.session_state.ocr_df = confirmed_df.copy()

            if brand == "AG":
                bad = confirmed_df["摘要"].map(lambda v: not str(v or "").strip().isdigit())
                if bad.any():
                    raise ValueError("AG 每一款都必須輸入純數字摘要。")

            if brand == "AN":
                final_df,working,code_by_key,statuses=assign_an_codes(
                    confirmed_df,ledger,custom_code,seq_mode,manual_starts
                )
            elif brand == "AG":
                final_df,working,code_by_key,statuses=assign_ag_codes(
                    confirmed_df,ledger,brand_code,custom_code,seq_mode,manual_starts,live_color_map
                )
            else:
                final_df,working,code_by_key,statuses=assign_jb_codes(
                    confirmed_df,ledger,brand_code,custom_code,seq_mode,jb_manual_sequences
                )

            purchase_outputs=create_purchase_workbooks(
                final_df,config_bytes["purchase_template"],brand,brand_code,optional_date,optional_page
            )
            master_name,master_bytes,master_df=create_master_workbook(
                final_df,config_bytes["master_template"],code_by_key,live_color_map,
                brand,brand_code,season,optional_date
            )

            zbio=io.BytesIO()
            with zipfile.ZipFile(zbio,"w",zipfile.ZIP_DEFLATED) as z:
                for name,data in purchase_outputs.items(): z.writestr(name,data)
                z.writestr(master_name,master_bytes)
            st.session_state.generated_zip=zbio.getvalue()

            # 永久保存流水號
            ledger_bytes=json.dumps(working,ensure_ascii=False,indent=2).encode("utf-8")
            gh_write_bytes(CONFIG_PATHS["ledger"],ledger_bytes,f"更新 {brand} 採購貨號流水")

            # 永久保存新增/更正廠商
            updated_vendor=dict(vendor_map); vendor_changed=False
            for _,r in final_df.iterrows():
                name=clean(r["廠商"]); code=str(r["廠商代碼"]).strip()
                if name and code and updated_vendor.get(name)!=code:
                    updated_vendor[name]=code; vendor_changed=True
            if vendor_changed:
                gh_write_bytes(CONFIG_PATHS["vendor"],vendor_map_to_excel(updated_vendor),f"更新 {brand} 廠商代碼")

            if brand in ("AN", "AG", "JB"):
                width = 3 if brand == "AN" else 4
                next_summary = "｜".join(
                    f'{cat} {int(working["next_sequence_by_category"].get(str(code),1)):0{width}d}'
                    for cat, code in CATEGORY_PRODUCT_CODE.items()
                    if cat in batch_categories
                )
                st.success(f"完成！{brand} 各類別流水號已永久保存｜{next_summary}")
            for s in statuses: st.write("•",s)

            st.subheader("本次採購單實際輸出資料")
            preview_cols = [
                c for c in ["廠商","原廠編號","貨號","類別","顏色","尺寸","數量","資料來源"]
                if c in final_df.columns
            ]
            st.dataframe(final_df[preview_cols], use_container_width=True, hide_index=True)

            st.subheader("商品基本資料預覽")
            st.dataframe(master_df,use_container_width=True)
        except Exception as e:
            st.error(f"無法產生：{e}")

    if st.session_state.generated_zip:
        st.download_button("📦 下載完整 ZIP",st.session_state.generated_zip,file_name=f"{brand}採購-完整輸出.zip",mime="application/zip",type="primary")

st.divider()
st.caption("三品牌永久保存版｜AN、AG、JB 固定設定與貨號流水分開保存。")
